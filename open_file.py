from openpyxl import load_workbook  # 파일 불러오기

wb = load_workbook("sample.xlsx")  # sample.xlsx 파일에서 wb 을 가져옴
ws = wb.active  # 활성화된 Sheet

# cell 데이터 불러오기
# for x in range(1, 11):  # 10 개 row
#     for y in range(1, 11):  # 10개 column
#         print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ...
#     print()  # 줄 바꿈

# cell 갯수를 모를 경우
for x in range(1, ws.max_row + 1):  # row
    for y in range(1, ws.max_column + 1):  # column
        print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ...
    print()
